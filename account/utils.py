import os
import openpyxl
from celery import shared_task
from .models import User, Loan
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from datetime import datetime

def export_user_to_excel(data):
    file_path = os.path.join(settings.MEDIA_ROOT, 'user_details.xlsx')
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    sheet.append(['user_id', 'first_name', 'last_name','phone_number','monthly_salary','approved_limit','current_debt'])

    # Write data
    for user in data:
        sheet.append([user.id, user.first_name, user.last_name, user.phone_number, user.monthly_salary, user.approved_limit, user.current_debt])

    workbook.save(file_path)

def credit_score(id):
    try:
        credit_score_val = 0
        try:
            user = User.objects.get(id=id)
        except ObjectDoesNotExist:
            raise ValueError("User not found")

        if user.loan_set.count() == 0:  # Accessing related Loan objects
            credit_score_val = 70
            return credit_score_val

        # Calculate past loans paid on time
        loans = Loan.objects.filter(user_id=id).all()
        tenure = 0
        for loan in loans:
            if loan.end_date < datetime.now().date():
                end_date = loan.end_date
            else:
                end_date = datetime.now().date()
            # Calculate tenure from start_date to end_date
            tenure = (end_date - loan.start_date).days / 30
            print(tenure)

        past_loans_paid_on_time = sum(loan.emi_paid for loan in user.loan_set.all())  # Accessing related Loan objects

        percent_past_loan_paid_on_time = round(((tenure - past_loans_paid_on_time) / tenure) * 100)
        emi_weight = 30 - (percent_past_loan_paid_on_time * 3) / 10

        # Calculate number of loans taken in the past
        number_of_loans_taken = user.loan_set.count()  # Accessing related Loan objects
        num_of_loan_weight = 10 - 10 / (number_of_loans_taken + 1)

        # Get current year
        current_year = datetime.now().year

        # Calculate loan activity in the current year
        current_year_loans = user.loan_set.filter(start_date__year=current_year) | user.loan_set.filter(end_date__year=current_year)  # Accessing related Loan objects

        current_year_loan_weight = 30 * (30 / 100000) * len(current_year_loans)

        # Calculate loan approved volume
        loan_approved_volume = user.approved_limit

        loan_approved_weight = max(0, min(30, (30 / 100000) * loan_approved_volume))

        # Get the sum of current loans from the current_debt field
        sum_current_loans = user.current_debt or 0
        # If the sum of all current EMIs > 50% of monthly salary, credit score = 0

        emis = sum(loan.emi for loan in user.loan_set.all())  # Accessing related Loan objects
        if emis > (50 / 100) * user.monthly_salary:
            credit_score_val = 0

        if sum_current_loans > loan_approved_volume:
            credit_score_val = 0
        else:
            credit_score_val = round(emi_weight + num_of_loan_weight + current_year_loan_weight + loan_approved_weight)

        return credit_score_val

    except Exception as e:
        raise ValueError(f"Error calculating credit score: {str(e)}")

#created a function import users & loan data from excel file in media folder and save to database
@shared_task
def import_loan_from_excel():
    file_path = os.path.join(settings.MEDIA_ROOT, 'loan_data.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        loan = Loan(
            user_id=row[0],
            loan_amount=row[2],
            tenure=row[3],
            interest_rate=row[4],
            emi=row[5],
            emi_paid=row[6],
            start_date=row[7],
            end_date=row[8]
        )
        loan.save()

@shared_task
def import_user_from_excel():
    file_path = os.path.join(settings.MEDIA_ROOT, 'customer_data.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user = User(
            first_name=row[1],
            last_name=row[2],
            age=row[3],
            phone_number=row[4],
            monthly_salary=row[5],
            approved_limit=row[6],
        )
        try:
            user.save()
        except Exception as e:
            print(f"Error saving user: {str(e)}")
            continue